"""
@author:maohui
@time:3/14/2023 1:21 PM
  　　　　　　　 ┏┓    ┏┓+ +
  　　　　　　　┏┛┻━━━━┛┻┓ + +
  　　　　　　　┃        ┃ 　 
  　　　　　　　┃     ━  ┃ ++ + + +
  　　　　　 　████━████ ┃+
  　　　　　　　┃        ┃ +
  　　　　　　　┃   ┻    ┃
  　　　　　　　┃        ┃ + +
  　　　　　　　┗━┓   ┏━━┛
  　　　　　　　  ┃   ┃
  　　　　　　　  ┃   ┃ + + + +
  　　　　　　　  ┃   ┃　　　Code is far away from bug with the animal protecting
  　　　　　　　  ┃   ┃+ 　　　　神兽保佑,代码无bug
  　　　　　　　  ┃   ┃
  　　　　　　　  ┃   ┃　　+
  　　　　　　　  ┃   ┗━━━━━━━┓ + +     
  　　　　　　　  ┃           ┣┓
  　　　　　　　  ┃           ┏┛
  　　　　　　　  ┗┓┓┏━━━━━┳┓┏┛ + + + +
  　　　　　　　   ┃┫┫     ┃┫┫
  　　　　　　　   ┗┻┛     ┗┻┛+ + + +
"""
import base64
import io
import os
import re
import urllib

import requests
from PIL import Image
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

# 文字识别
API_KEY = "gqDNeW5jSprM6qhpHi9Ay2W0"
SECRET_KEY = "neVQc8VMIzhVl4LrEa4EXW0k0F7mvdxf"


#
# API_KEY = "BmIGkZXpWWc7G89x9ZKklluo"
# SECRET_KEY = "dOGcGel3NXwISvdYAHGCeiWyKPoU6pvF"


def main():
    path = input("请输入图片文件夹绝对路径：")
    files_list = select_file(path)

    for file in files_list:
        try:
            # f = open(adjust_img_size(file), 'rb')
            # img = base64.b64encode(f.read())
            # print(img.__sizeof__())
            print(file)
            url = "https://aip.baidubce.com/rest/2.0/ocr/v1/accurate_basic?access_token=" + get_access_token()  # ocr高
            payload = f'image={get_file_content_as_base64(adjust_img_size(file), True)}&detect_direction=false&paragraph=false&probability=false'
            headers = {
                'Content-Type': 'application/x-www-form-urlencoded',
                'Accept': 'application/json'
            }

            response = requests.request("POST", url, headers=headers, data=payload)
            print(response.text)
            response_result = response.json().get("words_result")

            context_list = get_info(file, detect_datetime(response_result), response_result)
            output(context_list)
            os.remove(adjust_img_size(file))
        except:
            context_list = get_info(file, 0, 0)
            output(context_list)


def get_file_content_as_base64(path, urlencoded=False):
    """
    获取文件base64编码
    :param path: 文件路径
    :param urlencoded: 是否对结果进行urlencoded
    :return: base64编码信息
    """
    with open(path, "rb") as f:
        content = base64.b64encode(f.read()).decode("utf8")
        if urlencoded:
            content = urllib.parse.quote_plus(content)
    return content


def get_access_token():
    """
    使用 AK，SK 生成鉴权签名（Access Token）
    :return: access_token，或是None(如果错误)
    """
    url = "https://aip.baidubce.com/oauth/2.0/token"
    params = {"grant_type": "client_credentials", "client_id": API_KEY, "client_secret": SECRET_KEY}
    return str(requests.post(url, params=params).json().get("access_token"))


def detect_datetime(words_result):
    # result = {"words_result": [{"words": "分子编号"}, {"words": "灶石"}, {"words": "性别年龄"}, {"words": "登记日期"},
    #                            {"words": "送检材料"}, {"words": "检查类型"}, {"words": "报告类型"},
    #                            {"words": "报告状态"}, {"words": "备注"}, {"words": "延迟原因"}, {"words": "T20-3376"},
    #                            {"words": "1682909"}, {"words": "男"}, {"words": "54"}, {"words": "2020-12-30"},
    #                            {"words": "08:48:48"}, {"words": "特需会诊"}, {"words": "特需会诊"},
    #                            {"words": "已发报告"}, {"words": "1021023"}, {"words": "1682909"}, {"words": "男"},
    #                            {"words": "54"}, {"words": "2021-02-24"}, {"words": "15:09:58"},
    #                            {"words": "直肠及肛周组织"}, {"words": "常规"}, {"words": "常规"}, {"words": "已发报告"},
    #                            {"words": "免疫组化"}, {"words": "免疫组化"}, {"words": "常规：直肠及肛周组织"},
    #                            {"words": "检查日期"}, {"words": "2021-02-2415:09:58"}, {"words": "病理号/细胞学号"},
    #                            {"words": "1021023"}, {"words": "检查类型"}, {"words": "常规"}, {"words": "标本类型"},
    #                            {"words": "大标本"}, {"words": "送检材料"}, {"words": "直肠及肛周组织"},
    #                            {"words": "报告日期"}, {"words": "2021-03-0209:02:31"}, {"words": "申请科室"},
    #                            {"words": "结畅外科2"}, {"words": "申请医师"}, {"words": "报告医师"},
    #                            {"words": "审核医师"}, {"words": "主诊医师"}, {"words": "肉眼所见"}, {
    #                                "words": "(1)管一段，长17cm,上切缘宽4.5cm,下附肛周组织，皮肤面积6×8.5cm,距上切缘12cm,距齿权5x0.4cm,可见一盘状隆起型肿物，大小4.8×4.5×0.8cm,累及肌层，似及肠周脂肪，"},
    #                            {
    #                                "words": "距环周切缘最近2cm,未票及齿状线，于肠周脂肪组织中找到结节数枚，直径0.3-0.6cm。于肿物上方距肿物最近1.5cm肠周脂肪组织中可见一癌结节，直径2.2cm,界尚清。"},
    #                            {"words": "(1)(直肠及肛周组织)"}, {
    #                                "words": "直肠盘状隆起型中-低分化腺癌，肿瘤穿透固有肌层至肠周纤维脂肪组织，累及齿状线。可见神经侵犯，未见明确脉管瘤栓。上切缘、下切缘及环周切缘均未见癌。"},
    #                            {"words": "淋巴结见转移癌(1/16)"}, {"words": "诊断意见"}, {"words": "肠系膜淋巴结0/8"},
    #                            {"words": "肠壁淋巴结1/8"}, {"words": "pTNM分期：pT3N1a"}, {
    #                                "words": "免疫组化结果显示：BRAF-V600E(-),C-MET(2+),HER2(1+),MLH1(+),MSH2(+),MSH6(+), PMS2(+), Desmin(显示肌层)。"}],
    #           "words_result_num": 61, "log_id": 1722496941039673361}
    # words_result = result.get('words_result')
    for index, result in enumerate(words_result):
        res = re.search(r'[0-9]{4}[-/年]?[0-9]{2}[-,/月]?[0-9]{2}', result.get("words"))
        if res is not None:
            condition_png = "检查" in result.get("words") or "检查" in words_result[index - 1].get("words")
            condition_jpg = "收到" in result.get("words") or "收到" in words_result[index - 1].get("words")
            if condition_png or condition_jpg:
                print(result)
                return index, result
    return None, None


def adjust_img_size(image_path, target_quality=95):
    # img = Image.open(img_path)
    # img_bytes = io.BytesIO()
    # img.save(img_bytes, format='JPEG', quality=target_quality)  # 以JPEG格式保存图像到内存
    # img_size = img_bytes.tell()
    #
    # while img_size > 10 * 1024 * 1024:
    #     target_quality -= 5  # 逐渐降低压缩质量
    #     img_bytes = io.BytesIO()
    #     img.save(img_bytes, format='JPEG', quality=target_quality)
    #     img_size = img_bytes.tell()
    #
    # return img
    # filedir = os.path.join(os.path.dirname(image_path), "demo")
    # if not os.path.exists(filedir):
    #     os.makedirs(filedir)
    filename = os.path.basename(image_path).split('.')[0]
    filetype = os.path.basename(image_path).split('.')[1]
    output_filename = os.path.join(os.path.dirname(image_path), f'{filename}_1.{filetype}')
    check_image_orientation(image_path)
    with Image.open(image_path) as img:
        img.thumbnail((4096, 4096))
        # if filetype == "jpg":
        #     img.save(output_filename, "JPEG", quality=95)  # Adjust the quality as needed (95 is a good starting point)
        # if filetype == "png":
        img.save(output_filename, quality=95)  # Adjust the quality as needed (95 is a good starting point)
    # check_image_orientation(output_filename)
    return output_filename
    # imgSize = img.size  # 大小/尺寸
    # # print(imgSize)
    # w = img.width  # 图片的宽
    # h = img.height  # 图片的高
    # f = img.format  # 图像格式
    #
    # # if w > 4096 or h > 4096:
    # #  处理图片后存储路径，以及存储格式
    # img.save(f'{filename}_1.{filetype}')
    #
    # # #  resize图片大小，入口参数为一个tuple，新的图片的大小
    # # img_size = img.resize((w, h))
    # # print(img_size)
    #
    # return f'{filename}_1.{filetype}'


def select_file(filedir):
    # files = [os.path.join(filedir, file) for file in os.listdir(filedir)]
    # print(files)
    import os

    # dir_path = '当前目录'
    files_list = []
    dir_path = filedir
    for dirpath, dirnames, filenames in os.walk(dir_path):
        for filename in filenames:
            if (filename.endswith(".jpg") or filename.endswith(
                    ".png")) and ("cea" not in filename.lower()) and (not ((filename.lower()).endswith("ct"))):
                files_list.append(os.path.join(dirpath, filename))
                # print(os.path.join(dirpath, filename))
    return files_list


def check_image_orientation(image_path):
    """
    检查图片的朝向，并修改为垂直朝向
    @param image_path:
    @return:
    """

    # image_path = "1089704_1.jpg"
    # print(image_path)
    with Image.open(image_path) as img:
        try:
            exif_data = img._getexif()
            if exif_data is not None:
                orientation = img._getexif()[274]  # 274 corresponds to the EXIF orientation tag
                # print(orientation)
                if orientation == 3:
                    # The image is upside down. You may want to rotate it.
                    img = img.rotate(180, expand=True)
                    # return "Upside Down"
                elif orientation == 6:
                    # The image is rotated 90 degrees clockwise. You may want to rotate it.
                    img = img.rotate(270, expand=True)
                    # return "Rotated 90 degrees clockwise"
                elif orientation == 8:
                    # The image is rotated 90 degrees counterclockwise. You may want to rotate it.
                    img = img.rotate(90, expand=True)  # Rotate the image 180 degrees (upside down)
                    # return "Rotated 90 degrees counterclockwise"
                img.save(image_path)
        except (AttributeError, KeyError):
            # No EXIF orientation tag found, or it's not supported.
            img.save(image_path)
            # return "No EXIF orientation tag"


def get_info(file, detect_time, detect_res):
    context_list = []
    id = os.path.basename(file).split('.')[0]
    try:
        context_list.append(id)
        detect_index, detect_time = detect_time
        if detect_time is None:
            context_list = [id, "", ""]
        elif "收到" in detect_time.get("words") or "收到" in detect_res[detect_index - 1].get("words"):
            detect_time_digit = filter(lambda ch: ch in '0123456789', detect_time.get("words"))
            detect_time_digit_list = list(detect_time_digit)[0:8]
            date = "".join(detect_time_digit_list)
            context_list.append(date)
        elif "检查" in detect_time.get("words") or "检查" in detect_res[detect_index - 1].get("words"):
            context_list.append("")
            detect_time_digit = filter(lambda ch: ch in '0123456789', detect_time.get("words"))
            detect_time_digit_list = list(detect_time_digit)[0:8]
            date = "".join(detect_time_digit_list)
            context_list.append(date)
        else:
            context_list = [id, "", "", detect_time.get("words")]
    except:
        context_list = [id, "1", "1"]
    return context_list


def output(context_list):
    # filename = os.path.join(os.path.dirname(file), "results.xlsx")
    filename = "results.xlsx"
    if not os.path.exists(filename):
        workbook = Workbook()
        # worksheet = workbook.active
        workbook.save(filename)
    workbook = load_workbook(filename)
    worksheet = workbook.active
    worksheet.append(context_list)
    workbook.save(filename)


if __name__ == '__main__':
    # detect_datetime(1)
    main()
    # output(0, [0, 0, 0])
    # check_image_orientation("./demo/1156150.jpg")
    # compress_image("./demo/1156150.jpg")
    # get_file_content_as_base64("1181512-肠镜.jpg")
    # adjust_img_size("1181512-肠镜.jpg")
    # select_file("../png_ocr/demo")
